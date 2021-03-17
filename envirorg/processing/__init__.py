import tempfile
import os
import calliope
import pandas as pd
from typing import Dict, List, Tuple

from nexinfosys.command_generators import Issue
from nexinfosys.embedded_nis import NIS
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_workbook


class LCIIndex:
    """ An index of where is each activity, to ease reading it """

    def __init__(self, lci_data_locations: List[str]):
        # TODO Prepare an index of activities, and where each is stored. Maybe also the flows
        pass

    def activities(self):
        """
        All indexed activities

        :return:
        """

    def activity_information(self, activity):
        pass


class Simulation:
    """ Abstract simulation reader (the implementation will depend on the type of simulator used) """
    def read(self, simulation_files_path: str):
        """ Read"""
        pass

    def blocks(self):
        """ An iterator into the blocks (technologies) in the simulation """
        pass

    def block_information(self, block_name):
        """ Return information for block """
        pass


class CalliopeSimulation(Simulation):
    def __init__(self, simulation_files_path):
        self._model = None
        self.read(simulation_files_path)

    def read(self, simulation_files_path: str):
        self._model = calliope.Model(simulation_files_path)

    def blocks(self):
        """ List of technologies """

    def block_information(self, block_name):
        """ Return information for block """
        pass


class Matcher:
    """ * Match Technologies (simulation) into parent Processors (MuSIASEM)
        * Match Activities (LCA) into Technologies (simulation)
    """
    def __init__(self, correspondence_files_path):
        self.correspondence_location = correspondence_files_path
        self._read()

    def _read(self):
        """ Read the correspondence files """
        # TODO
        pass

    def find_related_lci_activities(self, block, lci_index: LCIIndex):
        # TODO
        pass

    def find_candidate_musiasem_parent_processors(self, nis: NIS, block):
        # TODO Find parent processors names in the correspondence, locate them in NIS and return the full hierarchy name
        #      of each
        pass

    def equivalent_interface_type_for_lci_flow(self, lca_flow: str, nis: NIS):
        pass

    def equivalent_interface_type_for_simulation_flow(self, simulation_flow: str, nis: NIS):
        pass


def read_parse_configuration(file_name) -> Dict:
    def read_yaml_configuration(file_content) -> Dict:
        import yaml
        d = yaml.load(file_content, Loader=yaml.FullLoader)
        return d

    # Read into string
    with open(file_name, 'r') as file:
        contents = file.read()

    # Read depending on format
    if file_name.endswith("yaml") or file_name.endswith("yml"):
        cfg = read_yaml_configuration(contents)

    # Check configuration
    mandatory_keys = ["nis_file_location",
                      "correspondence_files_path",
                      "simulation_type", "simulation_files_path",
                      "lci_data_locations",
                      "output_directory"]
    any_not_found = False
    for k in mandatory_keys:
        if k not in cfg:
            print(f"{k} key not found in the configuration file {file_name}")
            any_not_found = True
    if any_not_found:
        raise Exception("Configuration file incomplete or incorrect. See log for missing keys")
    return cfg


def read_prepare_nis_file(nis_file_url: str) -> Tuple[NIS, List[Issue]]:
    nis = NIS()
    nis.open_session(True)
    nis.reset_commands()
    # Load file and execute it
    nis.load_workbook(nis_file_url)
    issues = nis.submit_and_solve()
    # tmp = nis.query_available_datasets()
    # print(tmp)
    # d = nis.get_results([(tmp[0]["type"], tmp[0]["name"])])
    return nis, issues


# Output file generation

def list_to_dataframe(lst: List) -> pd.DataFrame:
    return pd.DataFrame(data=lst[1:], columns=lst[0])


def generate_import_commands_command(files: List[str]) -> pd.DataFrame:
    lst = [["Workbook", "Worksheets"]]
    for file in files:
        lst.append([file, ""])

    return list_to_dataframe(lst)


"""
PROCESSORS | BLOCKS | ACTIVITIES 
[
 {"name": "",
  "parent": "",
  "group": "",
  "subsystem_type": "local",
  "system": "",
  "function_or_structural": "Structural",
  "accounted": "yes",
  "parent_membership_weight": "",
  "behave_as": "",
  "level": "",
  "description": "",
  "geolocation_ref": "",
  "geolocation_code": "",
  "geolocation_lat_long": "",
  "attributes": {
   "att": "value",
  }
  "interfaces": [
    {"": "",
     "": "",
     "attributes": {
      "att": "value",
     }
     "observations": [
       {
        "": "",
        "attributes": {
         "att": "value"
        }
       }
     ]
  ]
]
"""

def generate_processors_command(blocks_as_processor_templates: List) -> pd.DataFrame:
    lst = [["ProcessorGroup",
            "Processor",
            "ParentProcessor",
            "SubsystemType",
            "System",
            "FunctionalOrStructural",
            "Accounted",
            "ParentProcessorWeight",
            "BehaveAs",
            "Level",
            "Description",
            "GeolocationRef",
            "GeolocationCode",
            "GeolocationLatLong",
            "Attributes"]]
    for b in blocks_as_processor_templates:
        t = [b.get("group", ""),
             b.get("name"),
             b.get("parent"),
             b.get("subsystem_type", ""),
             b.get("system", ""),
             b.get("functional_or_structural", "Structural"),
             b.get("accounted", "Yes"),
             b.get("parent_membership_weight", ""),
             b.get("behave_as", ""),
             b.get("level", ""),
             b.get("description", ""),
             b.get("geolocation_ref", ""),
             b.get("geolocation_code", ""),
             b.get("geolocation_lat_long", ""),
             b.get("attributes", ""),  # Transform into a comma separated list of key=value
             ]
        lst.append(t)

    return list_to_dataframe(lst)


def generate_interfaces_command(blocks_as_processor_templates: List) -> pd.DataFrame:
    lst = [[
        # Interface
        "Processor",
        "InterfaceType",
        "Interface",
        "Sphere",
        "RoegenType",
        "Orientation",
        "OppositeSubsystemType",
        "GeolocationRef",
        "GeolocationCode",
        "Range",
        "RangeUnit",
        "InterfaceAttributes",
        # Quantity
        "Value",
        "Unit",
        "RelativeTo",
        "Uncertainty",
        "Assessment",
        "PedigreeMatrix",
        "Pedigree",
        "Time",
        "Source",
        "NumberAttributes",
        "Comments"
    ]]
    for b in blocks_as_processor_templates:
        processor = f'{b.get("parent")}.{b.get("name")}'
        for i in b.get("interfaces"):
            t = [processor,
                 i.get("interface_type"),
                 i.get("interface"),
                 i.get("sphere", ""),
                 i.get("roegen_type", ""),
                 i.get("orientation"),
                 i.get("opposite_subsystem_type", ""),
                 i.get("geolocation_ref", ""),
                 i.get("geolocation_code", ""),
                 i.get("range", ""),
                 i.get("attributes", ""),  # Transform into a comma separated list of key=value
                 ]
            if len(i["observations"]) > 0:
                for o in i["observations"]:
                    s = [
                        o.get("value"),
                        o.get("unit", ""),
                        o.get("relative_to", ""),
                        o.get("uncertainty", ""),
                        o.get("assessment", ""),
                        o.get("pedigree_matrix", ""),
                        o.get("pedigree", ""),
                        o.get("time"),
                        o.get("source"),
                        o.get("attributes", ""),  # Transform into a comma separated list of key=value
                        o.get("comments", "")
                    ]
                    lst.append(t+s)
            else:
                lst.append(t+[""]*11)

    # Convert to pd.DataFrame
    return list_to_dataframe(lst)


def generate_workbook_from_list_of_worksheets(output_name: str, lst: List[Tuple[str, pd.DataFrame]]):
    """

    :param output_name: Name of the output file
    :param lst: List of tuples (worksheet name, worksheet contents)
    :return:
    """
    # Convert list of pd.DataFrames to Excel workbook
    wb = Workbook(write_only=True)
    for name, df in lst:
        if df.shape[0] < 2:
            continue

        ws = wb.create_sheet(name)
        widths = [0]*(df.shape[1]+1)  # A maximum of 100 columns
        max_columns = 0
        for r in dataframe_to_rows(df, index=False, header=True):
            if len(r) > max_columns:
                max_columns = len(r)
            for i in range(len(r)):
                width = int(len(str(r[i])) * 1.1)
                if width > widths[i]:
                    widths[i] = width

        for i, column_width in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    return save_workbook(wb, output_name)


def merge_models(nis: NIS, matcher: Matcher, simulation: Simulation, lci_data_index: LCIIndex):
    """
    Inputs:
      - Assessment configuration
      - Base MuSIASEM NIS file
      - LCI data
      - Simulation
        - Structure
        - Inputs
        - Outputs
      - Correspondence file
    :return:
    """
    # TODO - Parse all inputs
    #        - Base MuSIASEM: read, have model
    #        - Simulation structure
    #        - Correspondence file:
    #          - NIS: every "leaf" processor is in the file
    #          - Simulation: every technology is in the file, and can be assigned to at least one processor (more than one?)
    #          - LCA: every technology has an LCI
    #      - Configuration: * SPACE. single system or system per location or per country?
    #                       * TIME. Single NIS or multiple iterations
    #      - Merge NIS file with Simulation and LCI data
    #      - Compute indicators. Data structure: a dataset (dimensions - values - attributes) for each multilevel, another for scalar
    #      - Generate: jupyter script Python-R with example graphs. Network graph?, Sankey?, Geolayer?
    blocks_as_processor_templates = []
    for block in simulation.blocks:
        # Information (interfaces) of block from the simulation
        # TODO Name, value, unit, input/output, time, location, source, etc.
        block_info = simulation.read_block_info(block)
        # Matching LCA information
        matching_lci_activities = matcher.find_related_lci_activities(block, lci_data_index)
        for lci_activity in matching_lci_activities:
            for activity_flow in lci_activity.flows:
                # TODO Name, value, unit, input/output, time, location, source, etc.
                block_info.update()
        # TODO Convert input information into NIS InterfaceTypes (including change of scale?)
        block_info = matcher.convert_to_interface_types(block_info)
        # Matching MuSIASEM processors
        parents = matcher.find_candidate_musiasem_parent_processors(nis, block)
        # TODO obtain full processor names
        block_info["parents"] = parents
        # Append one more processor
        blocks_as_processor_templates.append(block_info)
    # Generate NIS commands
    lst = []
    lst.append(("Import", generate_import_commands_command([])))  # Input NIS file path
    # Enumerate the blocks and relate them to parent processors
    lst.append(("BareProcessors", generate_processors_command(blocks_as_processor_templates)))
    lst.append(("Interfaces", generate_interfaces_command(blocks_as_processor_templates)))
    # TODO Dump
    #  blocks with no MuSIASEM parents (as warning?),
    #  blocks with no matching LCI (as warning)
    #  flows with no matching InterfaceType (as error)
    return lst


def enviro_musiasem(cfg_file_path):
    """
    From the configuration file, read all inputs
      - Base NIS format file
      - Correspondence file
      - Simulation type, simulation location
      - LCI databases

    :param configuration_file:
    :return:
    """
    # Read configuration
    cfg = read_parse_configuration(cfg_file_path)

    # Construct auxiliary models
    # Matcher
    matcher = Matcher(cfg["correspondence_files_path"])
    # MuSIASEM (NIS)
    nis = read_prepare_nis_file(cfg["nis_file_location"])
    # LCI index
    lci_data_index = LCIIndex(cfg["lci_data_locations"])
    # Simulation
    if cfg["simulation_type"].lower() == "calliope":
        simulation = CalliopeSimulation(cfg["simulation_files_path"])

    # ELABORATE MERGE NIS FILE
    lst = merge_models(nis, matcher, simulation, lci_data_index)

    # Generate NIS file into a temporary file
    temp_name = tempfile.NamedTemporaryFile(dir=cfg["output_directory"], delete=False)
    generate_workbook_from_list_of_worksheets(temp_name, lst)

    # Execute the NIS file
    nis, issues = read_prepare_nis_file(temp_name)

    # TODO Download outputs, elaborate indicators, write indicator files

    # os.remove(temp_name)


enviro_musiasem("../../example_config.yaml")
