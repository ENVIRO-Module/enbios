import copy
import json
import os
from io import BytesIO
from typing import List, Union, BinaryIO
from zipfile import ZipFile

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook

JSON_INDENT = 4
ENSURE_ASCII = False


def _json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    from datetime import datetime

    if isinstance(obj, datetime):
        serial = obj.isoformat()
        return serial
    elif isinstance(obj, np.int64):
        return int(obj)
    elif getattr(obj, "Schema"):
        # Use "marshmallow"
        return getattr(obj, "Schema")().dump(obj)
    raise TypeError(f"Type {type(obj)} not serializable")


def generate_json(o):
    return json.dumps(o,
                      default=_json_serial,
                      sort_keys=True,
                      indent=JSON_INDENT,
                      ensure_ascii=ENSURE_ASCII,
                      separators=(',', ': ')
                      ) if o else None


def list_to_dataframe(lst: List) -> pd.DataFrame:
    return pd.DataFrame(data=lst[1:], columns=lst[0])


def get_scenario_name(prefix, s):
    def is_int(si):
        try:
            int(si)
            return True
        except ValueError:
            return False

    return f"{prefix}{s}" if is_int(s) else str(s)


def generate_workbook(cmds, generate_empty=False):
    # Convert list of pd.DataFrames to Excel workbook
    wb = Workbook(write_only=True)
    ws_count = 0
    for name, df in cmds:
        if df.shape[0] == 0:
            if not generate_empty:
                continue

        ws_count += 1
        ws = wb.create_sheet(name)
        widths = [0]*(df.shape[1]+1)  # A maximum of 100 columns
        max_columns = 0
        for r in dataframe_to_rows(df, index=False, header=True):
            if len(r) > max_columns:
                max_columns = len(r)
            for i in range(len(r)):
                width = int(len(str(r[i])) * 1.1)
                if width > widths[i]:
                    widths[i] = width

        for i, column_width in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    if ws_count > 0:
        return save_virtual_workbook(wb)
    else:
        return None


def hash_array(f: Union[str, bytes]):
    import hashlib
    m = hashlib.md5()
    if isinstance(f, str):
        m.update(f.encode("utf-8"))
    else:
        m.update(f)
    return m.digest()


def set_zip_timestamp(in_zip: Union[os.PathLike, str, BinaryIO], timestamp=(2020, 1, 1, 0, 0, 0)) -> BinaryIO:
    """
    Modify the timestamp of all files, to stabilize the hash

    :param in_zip: Zip file whose files timestamp will be modified
    :param timestamp: Tuple with the timestamp to set for all the files
    :return A BytesIO with the resulting Zip
    """
    out_zip = BytesIO()
    with ZipFile(in_zip, mode="r") as zin:
        with ZipFile(out_zip, mode="w") as zout:
            for zinfo in zin.infolist():
                data = zin.read(zinfo.filename)
                zinfo_new = copy.copy(zinfo)
                zinfo_new.date_time = timestamp
                zout.writestr(zinfo_new, data)
    return out_zip
