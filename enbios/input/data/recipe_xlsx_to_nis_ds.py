from io import StringIO

import numpy as np
import openpyxl
import pandas as pd

"""
* Different names same substance: remove name. Use CIRpy if no reference name found
"""


recipe_file = "/home/rnebot/GoogleDrive/AA_SENTINEL/workspace_20210331/ReCiPe2016_CFs_v1.1_20180117.xlsx"
workbook = openpyxl.load_workbook(recipe_file, data_only=True)
name_to_group = {"Global Warming": ("GWP", dict(name=1, alt_name= 2, formula=3, horizons=[4, 5, 6], row=4, unit_label_at=(2,), unit="kg")),
                 "Stratospheric ozone depletion": ("ODP", dict(name=1, alt_name=2, horizons=[3, 4, 5], row=5, unit_label_at=(2,), unit="kg")),
                 "Ionizing radiation": ("IRP", dict(name=1, alt_name=2, compartment=4, horizons=[5, 6, 7], row=4, unit_label_at=(2,), unit="kg")),
                 "Human damage ozone formation": ("HOFP", dict(name=2, compartment=4, horizons=[5, 6, 7], row=4, unit_label_at=(2,), unit="kg")),
                 "Particulate matter formation": ("PMFP", dict(name=2, horizons=[3, 4, 5], row=4, unit_label_at=(2,), unit="kg")),
                 "Ecosyste damage ozone formation": ("EOFP", dict(name=2, compartment=4, horizons=[5, 6, 7], row=4, unit_label_at=(2,), unit="kg")),
                 "Terrestrial acidification": ("AP", dict(name=1, formula=1, horizons=[4, 5, 6], row=5, unit_label_at=(2,), unit="kg")),
                 "Freshwater eutrophication": ("FEP", dict(name=1, compartment=2, formula=1, horizons=[4, 5, 6], row=5, unit_label_at=(2,), unit="kg")),
                 "Marine eutrophication": ("MEP", dict(name=1, compartment=2, formula=1, horizons=[4, 5, 6], row=5, unit_label_at=(2,), unit="kg")),

                 "Terrestrial ecotoxicity": ("ETPterrestrial", dict(name=2, compartment=4, formula=1, horizons=[5, 6, 7], row=4, unit_label_at=(2,), unit="kg")),
                 "Freshwater ecotoxicity": ("ETPfw", dict(name=2, compartment=5, formula=1, horizons=[6, 7, 8], row=4, unit_label_at=(2,), unit="kg")),
                 "Marine ecotoxicity": ("ETPmarine", dict(name=2, compartment=5, formula=1, horizons=[6, 7, 8], row=4, unit_label_at=(2,), unit="kg")),
                 "Human carcinogenic toxicity": ("CFhuman_mid_carc", dict(name=2, compartment=5, formula=1, horizons=[6, 7, 8], row=4, unit_label_at=(2,), unit="kg")),
                 "Human noncarcinogenic toxicity": ("CFhuman_mid_ncarc", dict(name=2, compartment=5, formula=1, horizons=[6, 7, 8], row=4, unit_label_at=(2,), unit="kg")),
                 # "Land transformation": ("LTx", dict(name=1, formula=1, horizons=[4, 5, 6], row=3, unit_label_at=None, unit="kg")),
                 # "Land occupation": ("LOcc", dict(name=1, formula=1, horizons=[4, 5, 6], row=5, unit_label_at=(2,), unit="kg")),
                 # "Water consumption": ("Wcons", dict(name=1, formula=1, horizons=[4, 5, 6], row=5, unit_label_at=(2,), unit="kg")),
                 "Mineral resource scarcity": ("MResScar", dict(name=1, formula=1, horizons=[3, 4, 5], row=4, unit_label_at=(2,), unit="kg")),
                 "Fossil resource scarcity": ("FFP", dict(name=1, formula=1, horizons=[3, 4, 5], row=4, unit_label_at=(2,), unit="kg"))
                 }

horizon_codes = ["I", "H", "E"]  # Individualist (20), Hierarchist (100), Egalitarian (>=1000)
# TODO "Reference name" should be Ecospold name...
reference_names = {}  # Reference names and list of alternative names.
alt_names = {}  # Alternative to reference names

_ = []  # Output
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    name = sheet.title.strip()
    if name not in name_to_group:
        continue
    print(name)
    tmp = name_to_group[name]
    ind = tmp[0]
    tmp = tmp[1]
    unit_label_at = tmp["unit_label_at"][0]
    unit = tmp["unit"]
    compartment_col = tmp.get("compartment")
    horizon_labels = [sheet.cell(row=unit_label_at, column=c).value for c in tmp["horizons"]]
    for r in range(tmp["row"], sheet.max_row):
        name = sheet.cell(row=r, column=tmp["name"]).value
        if name:
            name = name.strip()
        if not name:
            continue
        # TODO Find STANDARD name. Use CIRpy?
        name = name
        horizon_values = [sheet.cell(row=r, column=c).value for c in tmp["horizons"]]
        horizon_values = [h if h != "#N/A" else "" for h in horizon_values]
        compartment = sheet.cell(row=r, column=compartment_col).value if compartment_col else ""

        if horizon_values[0] is None or horizon_values[0] == "":
            continue
        for h, v in zip(horizon_codes, horizon_values):
            _.append(("ReCiPe2016", ind, name, unit, h, v, compartment))

df = pd.DataFrame(_, columns=["LCIAMethod", "LCIAIndicator", "Interface", "InterfaceUnit", "LCIAHorizon", "LCIACoefficient", "Compartment"])

s = StringIO()
s.write("# To regenerate this file, execute module 'recipe_xlsx_to_nis_ds.py'")
df.to_csv(s, index=False)
with open("/home/rnebot/Downloads/recipe2016.csv", "wt") as f:
    f.write(s.getvalue())


