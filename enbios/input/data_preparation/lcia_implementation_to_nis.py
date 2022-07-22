from io import StringIO
from typing import List

import openpyxl
import pandas as pd
from nexinfosys.command_generators.parser_ast_evaluators import get_nis_name


def convert_lcia_implementation_to_nis(lcia_implementation_file: str, lcia_file: str,
                                       method_like, method_is: List[str], include_obsolete, use_nis_name_syntax):
    def get_horizon(ind):
        return ""

    def adapt_name(name):
        return get_nis_name(name) if use_nis_name_syntax else name

    workbook = openpyxl.load_workbook(lcia_implementation_file, data_only=True)
    cfs = workbook["CFs"]
    units = workbook["units"] if "units" in workbook.sheetnames else workbook["Indicators"]
    methods = set()
    categories = set()
    indicators = {}  # And its unit
    horizons = set()
    compartments = {}  # Include subcompartments
    if method_like != "":
        method_like = method_like.lower()
    if method_is:
        method_is = [m.lower() for m in method_is]

    for r in range(2, units.max_row):
        method = units.cell(row=r, column=1).value.strip()
        if not include_obsolete:
            if "superseded" in method.lower() or "obsolete" in method.lower():
                continue
        if method_like != "":
            if method_like not in method.lower():
                continue
        if method_is is not None:
            if method.lower() not in method_is:
                continue

        category = units.cell(row=r, column=2).value.strip()
        indicator = units.cell(row=r, column=3).value.strip()
        unit = units.cell(row=r, column=4).value.strip()

        methods.add(method)
        categories.add(category)
        indicators[indicator] = unit

    print("METHODS")
    print(list(methods))
    print("INDICATORS")
    print(list(indicators.keys()))
    print("CATEGORIES")
    print(list(categories))
    _ = []  # Output
    for r in range(2, cfs.max_row):
        method = cfs.cell(row=r, column=1).value.strip()
        if method not in methods:
            continue
        category = cfs.cell(row=r, column=2).value.strip()
        indicator = cfs.cell(row=r, column=3).value.strip()
        # Guess horizon
        h = get_horizon(indicator)
        name = cfs.cell(row=r, column=4).value.strip()
        compartment = cfs.cell(row=r, column=5).value.strip()
        subcompartment = cfs.cell(row=r, column=6).value.strip()
        v = cfs.cell(row=r, column=7).value
        ind_unit = indicators[indicator]
        if not include_obsolete:
            if "superseded" in name.lower() or "obsolete" in name.lower():
                continue

        _.append((method,
                  category,
                  indicator,
                  ind_unit,
                  adapt_name(name),  # InterfaceType name
                  "",  # This assumes unit from the InterfaceType definition, which may not be correct
                  h,  # Horizon, which may be empty ("")
                  v,
                  compartment,
                  subcompartment)
                 )

    df = pd.DataFrame(_, columns=["LCIAMethod", "LCIACategory", "LCIAIndicator", "LCIAIndicatorUnit",
                                  "Interface", "InterfaceUnit",
                                  "LCIAHorizon", "LCIACoefficient",
                                  "Compartment", "Subcompartment"])

    s = StringIO()
    s.write(f"# To regenerate this file, execute 'enbios lcia_implementation_to_nis {lcia_implementation_file} {lcia_file} {method_like if method_like else ''} {'--include-obsolete' if include_obsolete else ''}'\n")
    df.to_csv(s, index=False)
    with open(lcia_file, "wt") as f:
        f.write(s.getvalue())


if __name__ == '__main__':
    convert_lcia_implementation_to_nis("/home/rnebot/GoogleDrive/AA_SENTINEL/LCIA_implementation_3.7.1.xlsx", "/home/rnebot/Downloads/test.csv")
