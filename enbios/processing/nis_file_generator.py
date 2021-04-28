import pandas as pd
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_workbook


def list_to_dataframe(lst: List) -> pd.DataFrame:
    return pd.DataFrame(data=lst[1:], columns=lst[0])


def generate_import_commands_command(files: List[str]) -> pd.DataFrame:
    lst = [["Workbook", "Worksheets"]]
    for file in files:
        lst.append([file, ""])

    return list_to_dataframe(lst)


"""
PROCESSORS | BLOCKS | ACTIVITIES 
[
 {"name": "",
  "parent": "",
  "group": "",
  "subsystem_type": "local",
  "system": "",
  "function_or_structural": "Structural",
  "accounted": "yes",
  "parent_membership_weight": "",
  "behave_as": "",
  "level": "",
  "description": "",
  "geolocation_ref": "",
  "geolocation_code": "",
  "geolocation_lat_long": "",
  "attributes": {
   "att": "value",
  }
  "interfaces": [
    {"": "",
     "": "",
     "attributes": {
      "att": "value",
     }
     "observations": [
       {
        "": "",
        "attributes": {
         "att": "value"
        }
       }
     ]
  ]
]
"""


def generate_processors_command(blocks_as_processor_templates: List) -> pd.DataFrame:
    lst = [["ProcessorGroup",
            "Processor",
            "ParentProcessor",
            "SubsystemType",
            "System",
            "FunctionalOrStructural",
            "Accounted",
            "ParentProcessorWeight",
            "BehaveAs",
            "Level",
            "Description",
            "GeolocationRef",
            "GeolocationCode",
            "GeolocationLatLong",
            "Attributes"]]
    for b in blocks_as_processor_templates:
        t = [b.get("group", ""),
             b.get("name"),
             b.get("parent"),
             b.get("subsystem_type", ""),
             b.get("system", ""),
             b.get("functional_or_structural", "Structural"),
             b.get("accounted", "Yes"),
             b.get("parent_membership_weight", ""),
             b.get("behave_as", ""),
             b.get("level", ""),
             b.get("description", ""),
             b.get("geolocation_ref", ""),
             b.get("geolocation_code", ""),
             b.get("geolocation_lat_long", ""),
             b.get("attributes", ""),  # Transform into a comma separated list of key=value
             ]
        lst.append(t)

    return list_to_dataframe(lst)


def generate_interfaces_command(blocks_as_processor_templates: List) -> pd.DataFrame:
    lst = [[
        # Interface
        "Processor",
        "InterfaceType",
        "Interface",
        "Sphere",
        "RoegenType",
        "Orientation",
        "OppositeSubsystemType",
        "GeolocationRef",
        "GeolocationCode",
        "Range",
        "RangeUnit",
        "InterfaceAttributes",
        # Quantity
        "Value",
        "Unit",
        "RelativeTo",
        "Uncertainty",
        "Assessment",
        "PedigreeMatrix",
        "Pedigree",
        "Time",
        "Source",
        "NumberAttributes",
        "Comments"
    ]]
    for b in blocks_as_processor_templates:
        processor = f'{b.get("parent")}.{b.get("name")}'
        for i in b.get("interfaces"):
            t = [processor,
                 i.get("interface_type"),
                 i.get("interface"),
                 i.get("sphere", ""),
                 i.get("roegen_type", ""),
                 i.get("orientation"),
                 i.get("opposite_subsystem_type", ""),
                 i.get("geolocation_ref", ""),
                 i.get("geolocation_code", ""),
                 i.get("range", ""),
                 i.get("attributes", ""),  # Transform into a comma separated list of key=value
                 ]
            if len(i["observations"]) > 0:
                for o in i["observations"]:
                    s = [
                        o.get("value"),
                        o.get("unit", ""),
                        o.get("relative_to", ""),
                        o.get("uncertainty", ""),
                        o.get("assessment", ""),
                        o.get("pedigree_matrix", ""),
                        o.get("pedigree", ""),
                        o.get("time"),
                        o.get("source"),
                        o.get("attributes", ""),  # Transform into a comma separated list of key=value
                        o.get("comments", "")
                    ]
                    lst.append(t+s)
            else:
                lst.append(t+[""]*11)

    # Convert to pd.DataFrame
    return list_to_dataframe(lst)


def generate_workbook_from_list_of_worksheets(output_name: str, lst: List[Tuple[str, pd.DataFrame]]):
    """

    :param output_name: Name of the output file
    :param lst: List of tuples (worksheet name, worksheet contents)
    :return:
    """
    # Convert list of pd.DataFrames to Excel workbook
    wb = Workbook(write_only=True)
    for name, df in lst:
        if df.shape[0] < 2:
            continue

        ws = wb.create_sheet(name)
        widths = [0]*(df.shape[1]+1)  # A maximum of 100 columns
        max_columns = 0
        for r in dataframe_to_rows(df, index=False, header=True):
            if len(r) > max_columns:
                max_columns = len(r)
            for i in range(len(r)):
                width = int(len(str(r[i])) * 1.1)
                if width > widths[i]:
                    widths[i] = width

        for i, column_width in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    return save_workbook(wb, output_name)
